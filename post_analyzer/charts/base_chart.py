# post_analyzer/charts/base_chart.py

"""
کلاس پایه برای تمام نمودارها.
- مدیریت شیت داده کمکی (مخفی)
- متدهای مشترک ساخت سری داده
"""

from abc import ABC, abstractmethod
from typing import List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.legend import Legend
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


class BaseChart(ABC):
    """
    پایه مشترک نمودارها.

    هر نمودار:
      1. یک شیت داده کمکی (hidden) می‌سازد
      2. نمودار openpyxl تولید می‌کند
      3. نمودار را در یک شیت مخصوص قرار می‌دهد
    """

    # ── ابعاد پیش‌فرض نمودار ──
    DEFAULT_WIDTH = 22
    DEFAULT_HEIGHT = 14

    def __init__(self, wb: Workbook, df: pd.DataFrame, sheet_prefix: str):
        self._wb = wb
        self._df = df
        self._data_sheet_name = f"_data_{sheet_prefix}"
        self._chart_sheet_name = sheet_prefix

    # ────────── شیت داده کمکی ──────────

    def _write_data_sheet(self, df_subset: pd.DataFrame) -> None:
        """DataFrame را به یک شیت کمکی (مخفی) می‌نویسد."""
        ws = self._wb.create_sheet(title=self._data_sheet_name)
        for r_idx, row in enumerate(
            dataframe_to_rows(df_subset, index=False, header=True), start=1
        ):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        ws.sheet_state = "hidden"

    def _get_data_sheet(self):
        return self._wb[self._data_sheet_name]

    # ────────── ساخت Reference ──────────

    def _make_ref(
        self, ws, min_col: int, max_col: int, min_row: int, max_row: int
    ) -> Reference:
        return Reference(
            ws,
            min_col=min_col,
            max_col=max_col,
            min_row=min_row,
            max_row=max_row,
        )

    def _make_cat_ref(self, ws, col: int, min_row: int, max_row: int) -> Reference:
        """Reference برای محور دسته‌بندی (X)"""
        return Reference(ws, min_col=col, max_col=col, min_row=min_row, max_row=max_row)

    # ────────── ساخت Series ──────────

    def _create_series(
        self,
        ws,
        data_col: int,
        min_row: int,
        max_row: int,
        title: str,
        color: Optional[str] = None,
    ) -> Series:
        """یک Series صحیح برای openpyxl می‌سازد."""
        ref = Reference(ws, min_col=data_col, min_row=min_row, max_row=max_row)
        series = Series(ref, title=title)
        if color:
            from openpyxl.chart.series import DataPoint
            from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
            series.graphicalProperties.solidFill = color
        return series

    # ────────── تنظیمات مشترک نمودار ──────────

    def _apply_common_settings(self, chart, title: str):
        """تنظیمات مشترک: عنوان، ابعاد، لجند"""
        chart.title = title
        chart.width = self.DEFAULT_WIDTH
        chart.height = self.DEFAULT_HEIGHT
        chart.style = 10

        # لجند کنار نمودار (سمت راست)
        chart.legend = Legend()
        chart.legend.position = "r"

    def _enable_data_labels(self, chart):
        """فعال‌سازی برچسب داده روی همه سری‌ها"""
        for s in chart.series:
            s.dLbls = DataLabelList()
            s.dLbls.showVal = True

    def _set_axis_labels(self, chart, x_title: str = "", y_title: str = ""):
        """تنظیم عنوان محورها"""
        if x_title and hasattr(chart, "x_axis"):
            chart.x_axis.title = x_title
            chart.x_axis.delete = False
            chart.x_axis.tickLblPos = "low"
        if y_title and hasattr(chart, "y_axis"):
            chart.y_axis.title = y_title
            chart.y_axis.delete = False
            chart.y_axis.tickLblPos = "low"

    # ────────── قرار دادن نمودار در شیت ──────────

    def _place_chart(self, chart, cell: str = "A1"):
        """نمودار را در شیت مخصوص قرار می‌دهد."""
        ws = self._wb.create_sheet(title=self._chart_sheet_name)
        ws.add_chart(chart, cell)

    # ────────── متد اصلی ──────────

    @abstractmethod
    def create(self) -> None:
        """نمودار را بساز و در شیت قرار بده."""
        ...
