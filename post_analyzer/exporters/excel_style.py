# post_analyzer/exporters/excel_style.py

"""
کلاس استایل مشترک برای تمام شیت‌های اکسل.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


class ExcelStyle:
    """استایل‌دهی یکپارچه به شیت‌های اکسل"""

    # ── رنگ‌ها ──
    HEADER_BG = "1F4E79"
    HEADER_FONT_COLOR = "FFFFFF"
    TOTAL_ROW_BG = "FFF2CC"
    EVEN_ROW_BG = "D6E4F0"
    ODD_ROW_BG = "FFFFFF"

    # ── فونت‌ها ──
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TOTAL_FONT = Font(bold=True, size=11)
    NORMAL_FONT = Font(size=10)

    # ── Fill ──
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    EVEN_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    ODD_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # ── Alignment ──
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Border ──
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── کلمات کلیدی برای فرمت‌دهی خودکار ──
    PERCENT_KEYWORDS = ['نسبت', 'درصد', '(%)']
    TIME_KEYWORDS = ['ساعت']

    @classmethod
    def apply_to_sheet(cls, ws: Worksheet, has_total_row: bool = False):
        """استایل کامل به یک شیت"""
        if ws.max_row is None or ws.max_row < 1:
            return

        max_col = ws.max_column
        max_row = ws.max_row

        cls._style_header(ws, max_col)
        cls._style_data_rows(ws, max_row, max_col, has_total_row)
        if has_total_row:
            cls._style_total_row(ws, max_row, max_col)
        cls._auto_column_width(ws, max_col, max_row)
        cls._apply_column_formats(ws, max_row, max_col)

    @classmethod
    def _style_header(cls, ws: Worksheet, max_col: int):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = cls.CENTER_ALIGN
            cell.border = cls.THIN_BORDER

    @classmethod
    def _style_data_rows(cls, ws: Worksheet, max_row: int, max_col: int, has_total_row: bool):
        last_data_row = (max_row - 1) if has_total_row else max_row
        for row in range(2, last_data_row + 1):
            is_even = (row % 2 == 0)
            fill = cls.EVEN_FILL if is_even else cls.ODD_FILL
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.font = cls.NORMAL_FONT
                cell.alignment = cls.CENTER_ALIGN
                cell.border = cls.THIN_BORDER

    @classmethod
    def _style_total_row(cls, ws: Worksheet, max_row: int, max_col: int):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=max_row, column=col)
            cell.fill = cls.TOTAL_FILL
            cell.font = cls.TOTAL_FONT
            cell.alignment = cls.CENTER_ALIGN
            cell.border = cls.THIN_BORDER

    @classmethod
    def _auto_column_width(cls, ws: Worksheet, max_col: int, max_row: int):
        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            max_len = 0
            for row in range(1, max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    cell_len = len(str(val))
                    if cell_len > max_len:
                        max_len = cell_len
            ws.column_dimensions[col_letter].width = max(min(max_len + 4, 50), 10)

    @classmethod
    def _apply_column_formats(cls, ws: Worksheet, max_row: int, max_col: int):
        for col in range(1, max_col + 1):
            header_val = str(ws.cell(row=1, column=col).value or '')
            is_percent = any(kw in header_val for kw in cls.PERCENT_KEYWORDS)
            is_time = any(kw in header_val for kw in cls.TIME_KEYWORDS)

            if is_percent:
                for row in range(2, max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.number_format = '0.00'
                    cell.alignment = cls.CENTER_ALIGN
            elif is_time:
                for row in range(2, max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = cls.CENTER_ALIGN
