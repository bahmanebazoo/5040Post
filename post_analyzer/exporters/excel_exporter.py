# post_analyzer/exporters/excel_exporter.py

"""
خروجی اکسل با استایل و نمودار.
"""

import pandas as pd
from typing import Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .base_exporter import BaseExporter
from .excel_style import ExcelStyle


class ExcelExporter(BaseExporter):
    """ذخیره DataFrameها + استایل + نمودار"""

    TOTAL_ROW_SHEETS = {'عملکرد اپراتورها'}
    VERTICAL_SHEETS = {'عملکرد تک‌روزه'}

    def __init__(self, output_file: str):
        self._output_file = output_file

    def export(
        self,
        sheets: Dict[str, pd.DataFrame],
        df_daily: Optional[pd.DataFrame] = None,
        df_operators: Optional[pd.DataFrame] = None,
    ) -> None:
        with pd.ExcelWriter(self._output_file, engine='openpyxl') as writer:
            # ── 1) نوشتن داده‌ها ──
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            # ── 2) اعمال استایل ──
            for sheet_name in list(writer.sheets.keys()):
                ws = writer.sheets[sheet_name]

                if sheet_name in self.VERTICAL_SHEETS:
                    self._style_vertical_sheet(ws)
                    print(f"  [Style] ✅ {sheet_name} (عمودی)")
                else:
                    has_total = sheet_name in self.TOTAL_ROW_SHEETS
                    ExcelStyle.apply_to_sheet(ws, has_total_row=has_total)
                    print(f"  [Style] ✅ {sheet_name}")
            for sheet_name in writer.sheets:
                writer.sheets[sheet_name].sheet_view.rightToLeft = True
                print(f"  [RTL] ✅ {sheet_name}")

            # ── 3) نمودارها ──
            if df_daily is not None and df_operators is not None:
                wb = writer.book
                print("\n[Charts] در حال ساخت نمودارها...")
                try:
                    from post_analyzer.charts.chart_manager import ChartManager
                    chart_manager = ChartManager(wb, df_daily, df_operators)
                    chart_manager.create_all_charts()

                    for ws in wb.worksheets:
                        ws.sheet_view.rightToLeft = True

                except Exception as e:
                    print(f"  [Charts] ❌ خطا در ساخت نمودارها: {e}")
                    import traceback
                    traceback.print_exc()

        print(f"\n[ExcelExporter] ✅ فایل ذخیره شد: {self._output_file}")

    @staticmethod
    def _style_vertical_sheet(ws) -> None:
        """استایل ویژه برای شیت عمودی (تک‌روزه)"""

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # --- رنگ‌ها ---
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(name='B Nazanin', size=12, bold=True, color='FFFFFF')

        title_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        title_font = Font(name='B Nazanin', size=13, bold=True, color='FFFFFF')

        param_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        param_font = Font(name='B Nazanin', size=11, bold=True, color='1F4E79')

        even_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
        odd_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        data_font = Font(name='B Nazanin', size=11, color='333333')

        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)

        max_row = ws.max_row
        max_col = ws.max_column

        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_align

                if row_idx == 1:
                    # ── ردیف هدر (نام ستون‌ها: پارامتر، اپراتور1، ...) ──
                    cell.fill = header_fill
                    cell.font = header_font

                elif row_idx == 2:
                    # ── ردیف عنوان تاریخ ──
                    cell.fill = title_fill
                    cell.font = title_font

                elif col_idx == 1:
                    # ── ستون اول (نام پارامترها) ──
                    cell.fill = param_fill
                    cell.font = param_font
                    cell.alignment = right_align

                else:
                    # ── سلول‌های داده ──
                    cell.font = data_font
                    if row_idx % 2 == 0:
                        cell.fill = even_fill
                    else:
                        cell.fill = odd_fill

        # --- تنظیم عرض ستون‌ها ---
        ws.column_dimensions['A'].width = 28  # ستون پارامتر
        for col_idx in range(2, max_col + 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = 22

        # --- ارتفاع ردیف‌ها ---
        ws.row_dimensions[1].height = 30  # هدر
        ws.row_dimensions[2].height = 35  # عنوان تاریخ
        for row_idx in range(3, max_row + 1):
            ws.row_dimensions[row_idx].height = 25

        # --- راست‌چین کل شیت (RTL) ---
        ws.sheet_view.rightToLeft = True
